"""OpenPyxl-based Spread Reader for extracting the target schema."""

from __future__ import annotations

from pathlib import Path
from typing import Any
import openpyxl
from decimal import Decimal


class SpreadReader:
    """Reads the target Excel proxy spread to generate a schema for the mapper."""

    def __init__(self, file_path: str | Path, sheet_name: str | None = None):
        self.file_path = Path(file_path)
        self.sheet_name = sheet_name

    def extract_schema(self, label_col: str = "B", prior_val_col: str | None = None, start_row: int = 1, end_row: int = 1000) -> list[dict[str, Any]]:
        """
        Scans the spread and builds a target schema list of dicts:
        [{'row': 5, 'label': 'CMV Total', 'prior_value': Decimal('100.5')}, ...]
        """
        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        
        if self.sheet_name and self.sheet_name in wb.sheetnames:
            ws = wb[self.sheet_name]
        else:
            ws = wb.active

        schema = []
        for row in range(start_row, end_row + 1):
            label_cell = ws[f"{label_col}{row}"]
            label_val = label_cell.value

            if not isinstance(label_val, str) or not label_val.strip():
                continue

            target_info: dict[str, Any] = {
                "row": row,
                "label": label_val.strip()
            }

            if prior_val_col:
                prior_cell = ws[f"{prior_val_col}{row}"]
                val = prior_cell.value
                if val is not None and isinstance(val, (int, float)):
                    target_info["prior_value"] = Decimal(str(val))

            schema.append(target_info)

        wb.close()
        return schema
