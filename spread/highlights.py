"""Highlights application for visual status tracking."""

from __future__ import annotations

from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill

from core.models import MappingResult


class Highlights:
    """Applies visual formatting to Excel files based on confidence levels."""

    def __init__(self, file_path: str | Path, sheet_name: str | None = None):
        self.file_path = Path(file_path)
        self.sheet_name = sheet_name

    def apply_styles(self, col: str, results: list[MappingResult]) -> None:
        """
        Highlights mapped results cells in Excel:
        - Green: confidence >= 0.95
        - Yellow: 0.60 <= confidence < 0.95
        """
        wb = openpyxl.load_workbook(self.file_path)
        
        if self.sheet_name and self.sheet_name in wb.sheetnames:
            ws = wb[self.sheet_name]
        else:
            ws = wb.active

        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for res in results:
            row = res.spread_row
            if row is None or res.value is None:
                continue

            cell = ws[f"{col}{row}"]

            conf = res.confidence
            if conf >= 0.95:
                cell.fill = green_fill
            elif 0.60 <= conf < 0.95:
                cell.fill = yellow_fill

        wb.save(self.file_path)
        wb.close()
