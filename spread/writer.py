"""OpenPyxl-based Spread Writer for populating target SpreadProxy."""

from __future__ import annotations

from pathlib import Path
import openpyxl

from core.models import MappingResult


class SpreadWriter:
    """Writes mapped results to an openpyxl Workbook."""

    def __init__(self, file_path: str | Path, sheet_name: str | None = None):
        self.file_path = Path(file_path)
        self.sheet_name = sheet_name

    def write_results(self, dest_col: str, results: list[MappingResult], overwrite: bool = True) -> None:
        """
        Dumps the values from MappingResult directly into the specified column of the Excel file.
        """
        wb = openpyxl.load_workbook(self.file_path)
        
        if self.sheet_name and self.sheet_name in wb.sheetnames:
            ws = wb[self.sheet_name]
        else:
            ws = wb.active

        for res in results:
            row = res.spread_row
            if not row or res.value is None:
                continue

            cell = ws[f"{dest_col}{row}"]

            # Only overwrite if specified, or if cell is empty
            if overwrite or cell.value is None:
                # We save it as float so Excel processes it natively as a number
                # instead of text
                cell.value = float(res.value)
                
            # Embed confidence as a comment or invisible trace if needed?
            # Or leave it for highs.py

        wb.save(self.file_path)
        wb.close()
