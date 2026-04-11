# processing/pipeline.py — Orquestrador principal do pipeline de Spread
# v2: corrigido motor duplo, auto-detecção de colunas, relatório melhor, multi-período

from __future__ import annotations

import logging
from pathlib import Path
from typing import Callable, List, Set, Tuple

import pandas as pd
from openpyxl import load_workbook

from core.utils import XLWINGS, periodos, col_txt_to_idx
from processing.origin import prepara_origem
from processing.spread import atualizar_ws, coletar_vals_do_spread
from processing.dre import aplicar_dre_manual
from processing.dfc import inserir_depreciacao_dfc
from processing.dmpl import inserir_dividendos_dm, inserir_aumentos_capital_dm
from processing.highlights import destacar_inseridos, destacar_novos
from processing.runtime_bridge import (
    next_data_column,
    special_row,
    spread_schema,
)

if XLWINGS:
    import xlwings as xw


def detectar_colunas(
    spread_path: Path,
    start_row: int | None = None,
    trimestral: bool = False,
) -> Tuple[str, str]:
    """
    Encontra a última coluna de dados preenchida no Spread e retorna
    (coluna_fonte, coluna_destino) respeitando o grid fixo D/F/H/J/L.

    Para períodos anuais, busca entre D/F/H/J.
    Para trimestral, a coluna destino é sempre L.
    """
    return detectar_colunas_normalized(
        spread_path,
        start_row=start_row,
        trimestral=trimestral,
    ).as_pair()


def detectar_colunas_normalized(
    spread_path: Path,
    start_row: int | None = None,
    trimestral: bool = False,
):
    """
    Retorna um resultado estruturado de auto-deteccao para consumo pela engine/UI.

    O wrapper legado detectar_colunas() continua retornando tupla apenas quando
    existe um slot livre explicito no grid configurado.
    """
    from engine.slot_detection import detect_spread_slot

    return detect_spread_slot(
        spread_path=spread_path,
        start_row=start_row,
        quarterly=trimestral,
    )


# ---------------------------------------------------------------------------
# Pipeline de período único
# ---------------------------------------------------------------------------
def _executar_pipeline_openpyxl(
    spr: Path,
    orig_abas: dict,
    df_dre, df_dfc, df_dm,
    src_idx: int, dst_idx: int,
    atual: str, ant: str,
    is_trim: bool,
    start_row: int,
    log: Callable[[str], None],
) -> Tuple[Path, Set[int], List[int]]:
    """Executa o pipeline via openpyxl. Retorna (output_path, used_vals, skipped_rows)."""
    is_xlsm = spr.suffix.lower() == ".xlsm"
    wb = load_workbook(spr, keep_vba=is_xlsm)
    schema = spread_schema()
    ws = wb[schema.sheet_name] if schema.sheet_name in wb.sheetnames else wb.active

    used_vals: Set[int] = set()

    skipped_rows, _, used = atualizar_ws(
        ws,
        lambda r, c: ws.cell(r, c).value,
        lambda r, c, v: setattr(ws.cell(r, c), "value", v),
        orig_abas, src_idx, dst_idx, atual, ant, start_row
    )
    used_vals |= used

    if is_trim and df_dre is not None:
        aplicar_dre_manual(df_dre, ws, dst_idx + 1, start_row, atual, False)

    if df_dfc is not None:
        amortizacao_row = special_row("amortizacao_total")
        if valor := inserir_depreciacao_dfc(
            df_dfc, ws, dst_idx + 1, amortizacao_row, atual, False
        ):
            used_vals.add(valor)

    if df_dm is not None:
        col_dm = dst_idx + 1
        neg, pos = inserir_dividendos_dm(
            df_dm,
            ws,
            col_dm,
            special_row("dividendos_pagos_negativo"),
            special_row("dividendos_pagos_positivo"),
            False,
        )
        if neg is not None:
            used_vals.add(neg)
        if pos is not None:
            used_vals.add(pos)

    if df_dm is not None:
        col_dm = dst_idx + 1
        if valor := inserir_aumentos_capital_dm(
            df_dm,
            ws,
            col_dm,
            special_row("reavaliacao_diferido"),
            False,
        ):
            used_vals.add(valor)

    out_name = f"{spr.stem} {atual}{'.xlsm' if is_xlsm else '.xlsx'}"
    out_path = spr.with_name(out_name)
    wb.save(out_path)
    wb.close()

    return out_path, used_vals, skipped_rows


def _executar_pipeline_xlwings(
    spr: Path,
    orig_abas: dict,
    df_dre, df_dfc, df_dm,
    src_idx: int, dst_idx: int,
    atual: str, ant: str,
    is_trim: bool,
    start_row: int,
    log: Callable[[str], None],
) -> Tuple[Set[int], List[int]]:
    """Executa o pipeline via xlwings. Retorna (used_vals, skipped_rows)."""
    for bk in xw.books:
        if Path(bk.fullname).resolve() == spr.resolve():
            wb = bk
            break
    else:
        wb = xw.Book(str(spr))

    schema = spread_schema()
    nomes = [s.name for s in wb.sheets]
    sht = wb.sheets[schema.sheet_name] if schema.sheet_name in nomes else wb.sheets.active

    get_val = lambda r, c: sht.cells(r, c).formula or sht.cells(r, c).value

    def set_val(r, c, v):
        attr = "formula" if isinstance(v, str) and v.startswith("=") else "value"
        setattr(sht.cells(r, c), attr, v)

    used_vals: Set[int] = set()

    skipped_rows, _, used = atualizar_ws(
        sht, get_val, set_val, orig_abas,
        src_idx, dst_idx, atual, ant, start_row
    )
    used_vals |= used

    if is_trim and df_dre is not None:
        aplicar_dre_manual(df_dre, sht, dst_idx + 1, start_row, atual, True)

    if df_dfc is not None:
        amortizacao_row = special_row("amortizacao_total")
        if valor := inserir_depreciacao_dfc(
            df_dfc, sht, dst_idx + 1, amortizacao_row, atual, True
        ):
            used_vals.add(valor)

    if df_dm is not None:
        col_dm = dst_idx + 1
        neg, pos = inserir_dividendos_dm(
            df_dm,
            sht,
            col_dm,
            special_row("dividendos_pagos_negativo"),
            special_row("dividendos_pagos_positivo"),
            True,
        )
        if neg is not None:
            used_vals.add(neg)
        if pos is not None:
            used_vals.add(pos)

    if df_dm is not None:
        col_dm = dst_idx + 1
        if valor := inserir_aumentos_capital_dm(
            df_dm,
            sht,
            col_dm,
            special_row("reavaliacao_diferido"),
            True,
        ):
            used_vals.add(valor)

    wb.app.calculate()
    wb.save()

    return used_vals, skipped_rows


# ---------------------------------------------------------------------------
# Função principal: 1 período
# ---------------------------------------------------------------------------
def processar(
    ori: Path,
    spr: Path,
    tipo: str,
    periodo: str,
    src_txt: str,
    dst_txt: str,
    start_row: int,
    out_dir: Path | None = None,
    log: Callable[[str], None] = print,
) -> Path:
    """Pipeline principal para um único período."""
    src_idx = col_txt_to_idx(src_txt)
    dst_idx = col_txt_to_idx(dst_txt)
    atual, ant, ant2, is_trim = periodos(periodo)

    log(f"Período: {atual} (anterior: {ant}) | Colunas: {src_txt}→{dst_txt}")

    # prepara e lê as abas tratadas
    orig_path = prepara_origem(ori, tipo, atual, ant, ant2, is_trim, out_dir)
    abas = pd.read_excel(orig_path, sheet_name=None, engine="openpyxl")

    prefix = 'cons' if tipo == 'consolidado' else 'ind'
    orig_abas = {
        f"{prefix} ativos": abas[f"{prefix} ativos"],
        f"{prefix} passivos": abas[f"{prefix} passivos"],
        f"{prefix} DRE": abas[f"{prefix} DRE"],
    }

    df_dre = abas.get(f"{prefix} DRE")
    df_dfc = abas.get(f"{prefix} DFC")
    df_dm = abas.get(f"{prefix} DMPL")

    # --- tenta xlwings primeiro; se falhar, usa openpyxl ---
    xlwings_ok = False
    used_vals: Set[int] = set()
    skipped_rows: List[int] = []

    if XLWINGS and spr.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            used, skipped = _executar_pipeline_xlwings(
                spr, orig_abas, df_dre, df_dfc, df_dm,
                src_idx, dst_idx, atual, ant, is_trim,
                start_row, log
            )
            used_vals |= used
            skipped_rows = skipped
            xlwings_ok = True
            log("✅ Processado via xlwings (Excel ao vivo)")
        except Exception as exc:
            log(f"⚠️ xlwings falhou ({exc}), usando openpyxl...")

    if not xlwings_ok:
        out_path, used, skipped = _executar_pipeline_openpyxl(
            spr, orig_abas, df_dre, df_dfc, df_dm,
            src_idx, dst_idx, atual, ant, is_trim,
            start_row, log
        )
        used_vals |= used
        skipped_rows = skipped
        spr = out_path
        log(f"✅ Processado via openpyxl → {spr.name}")

    # --- relatório ---
    _relatorio(spr, orig_path, dst_idx, start_row, atual, ant,
               used_vals, skipped_rows, log, prefer_xlwings=xlwings_ok)

    return spr


# ---------------------------------------------------------------------------
# Função multi-período: 3 períodos de uma vez
# ---------------------------------------------------------------------------
def processar_multi(
    ori: Path,
    spr: Path,
    tipo: str,
    periodo: str,
    src_txt: str,
    start_row: int,
    out_dir: Path | None = None,
    log: Callable[[str], None] = print,
) -> Path:
    """
    Processa 3 períodos de uma vez (a Origem CVM já tem 3 anos de dados).

    A partir da coluna-fonte (src_txt), preenche 3 colunas consecutivas:
      src → src+1 (ant2 → ant)
      src+1 → src+2 (ant → atual)
      ... ou seja, preenche da mais antiga para a mais recente.

    Retorna o path do Spread final.
    """
    atual, ant, ant2, is_trim = periodos(periodo)

    log(f"═══ Multi-período: {ant2} → {ant} → {atual} ═══")

    src_idx = col_txt_to_idx(src_txt)

    # Período 1: ant (mais antigo dos 2) — src → src+1 no grid fixo D/F/H/J/L
    # col_txt_to_idx é 0-based; get_column_letter é 1-based; o grid pula 1 coluna
    # oculta a cada passo, então o delta real é +3 (0-based +1 para 1-based +2 skip)
    dst1 = next_data_column(src_txt, include_quarterly=is_trim)
    log(f"\n── Período 1/2: {ant} (col {src_txt} → {dst1}) ──")

    # Para o primeiro período, precisamos de um período onde ant2 seria o "atual"
    # Usamos ant2 = periodo-2 como "periodo" → gera (ant2, ant3, ant4) mas só usamos ant2
    # Na verdade, é mais simples: rodamos cada período individualmente com os dados certos

    # A Origem já tem as 3 colunas renomeadas: atual, ant, ant2
    # Para multiperiodo, precisamos rodar 3 vezes trocando qual é a "fonte/destino"
    # na varredura do Spread

    # Mas o matching funciona assim: pega valor da col-fonte no Spread,
    # procura na Origem coluna "ant", retorna da Origem coluna "atual".
    # Então para preencher os 3 períodos, precisamos re-preparar a Origem
    # com diferentes pares de colunas como ant/atual.

    # Solução: rodar processar() 3 vezes com diferentes parâmetros de período
    # 1o: periodo que dá ant2 como ant e ant como atual => não existe diretamente
    # Mas podemos simplesmente rodar o pipeline 3 vezes incrementalmente.

    # Forma mais simples: processar 3 vezes em sequência
    # O período define quais colunas da Origem são "ant" e "atual"

    # Para 3 colunas, precisamos de 2 períodos intermediários:
    # - Primeiro preenche col_src+1 usando periodo=(ant) e src=src_txt
    # - Depois preenche col_src+2 usando periodo=(atual) e src=col_src+1

    # Determinar os 2 períodos intermediários
    if is_trim:
        # trimestral: ant = "nT(yy-1)", atual = "nTyy"
        # Para preencher 2 colunas, rodamos periodo=ant (gera ant como atual, ant2 como ant)
        # e depois periodo=atual (gera atual como atual, ant como ant)
        periodos_lista = [ant, atual]
    else:
        # anual
        periodos_lista = [ant, atual]

    resultado = spr
    col_atual = src_txt
    for i, per in enumerate(periodos_lista, 1):
        per_atual, per_ant, _, per_is_trim = periodos(per)
        dst_letter = next_data_column(col_atual, include_quarterly=per_is_trim)
        log(f"\n── Período {i}/{len(periodos_lista)}: {per_atual} (col {col_atual} → {dst_letter}) ──")
        resultado = processar(
            ori=ori, spr=resultado, tipo=tipo, periodo=per,
            src_txt=col_atual, dst_txt=dst_letter,
            start_row=start_row,
            out_dir=out_dir, log=log
        )
        col_atual = dst_letter

    log(f"\n═══ Multi-período concluído: {resultado.name} ═══")
    return resultado


# ---------------------------------------------------------------------------
# Relatório final
# ---------------------------------------------------------------------------
def _relatorio(
    spr: Path, orig_path: Path,
    dst_idx: int, start_row: int,
    atual: str, ant: str,
    used_vals: Set[int], skipped_rows: List[int],
    log: Callable[[str], None],
    prefer_xlwings: bool = False,
) -> None:
    """Gera destaques visuais e imprime relatório resumido."""
    spread_vals = coletar_vals_do_spread(spr, dst_idx, start_row)
    highlight = used_vals.union(spread_vals)
    destacar_inseridos(orig_path, highlight, atual, prefer_xlwings=prefer_xlwings)
    destacar_novos(orig_path, ant, atual, prefer_xlwings=prefer_xlwings)

    total_matched = len(used_vals)
    total_skipped = len(skipped_rows)
    missing = spread_vals - used_vals

    log(f"\n📊 Relatório:")
    log(f"  Valores correspondidos: {total_matched}")
    if total_skipped:
        log(f"  Linhas não correspondidas: {total_skipped} (linhas: {skipped_rows[:10]}{'...' if len(skipped_rows)>10 else ''})")
    if missing:
        log(f"  ⚠️ {len(missing)} valores no Spread sem correspondência na Origem")
    log(f"  Origem tratada: {orig_path.name}")
