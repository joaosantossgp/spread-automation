# processing/spread.py — Atualização da planilha Spread (varredura de valores/fórmulas)
# Extraído de app_spread.py
#
# MATCHING HÍBRIDO (dois estágios):
# ─────────────────────────────────
# Camada 2 (valor numérico, prioritária): pega o número na col-fonte do Spread, procura
#   na coluna `ant` da Origem, retorna o valor correspondente em `atual`. Comportamento
#   original. Pode ter falsos positivos quando valores coincidem entre contas distintas.
#
# Camada 1 (código CVM, fallback): lê o rótulo configurado no SpreadSchema,
#   resolve os CD_CONTA via MappingRegistry e soma os valores da coluna `atual`
#   nas abas da Origem. Determinístico, sem ambiguidade.

from __future__ import annotations

import re
from pathlib import Path
from typing import Callable, Dict, List, Set

import pandas as pd
from openpyxl import load_workbook

from core.utils import normaliza_num, adjust_complex_formula, shift_formula
from processing.runtime_bridge import (
    label_column_1based,
    layer1_codes_for_label,
    skip_rows,
    spread_sheet_name,
)


# ---------------------------------------------------------------------------
# Camada 2: matching por valor numérico (comportamento original)
# ---------------------------------------------------------------------------
def valor_corresp(
    abas: Dict[str, pd.DataFrame],
    n: int, prev: str, curr: str
) -> int | None:
    """
    Versão legada/individual do matching por valor.
    Otimizada para cachear índices no DataFrame (resolve o padrão de N+1 queries).
    """
    for df in abas.values():
        if prev not in df.columns or curr not in df.columns:
            continue

        cache_key = f"_indexed_{prev}"
        if cache_key not in df.attrs:
            norm_series = df[prev].apply(normaliza_num)
            df.attrs[cache_key] = df.set_index(norm_series)

        indexed_df = df.attrs[cache_key]
        if n in indexed_df.index:
            match = indexed_df.loc[n]
            if isinstance(match, pd.DataFrame):
                return normaliza_num(match[curr].iloc[0])
            return normaliza_num(match[curr])

    return None


def criar_mapa_corresp(
    abas: Dict[str, pd.DataFrame],
    prev: str, curr: str
) -> Dict[int, int]:
    """
    Cria um dicionário de lookup {valor_antigo: valor_atual} normalizado.
    Varre todas as abas da Origem.
    """
    mapa: Dict[int, int] = {}
    for df in abas.values():
        if prev not in df.columns or curr not in df.columns:
            continue
        # Extrai as colunas e normaliza em lote
        # Usamos dropna para ignorar valores que não normalizam para int
        temp_df = pd.DataFrame({
            'v_prev': df[prev].apply(normaliza_num),
            'v_curr': df[curr].apply(normaliza_num)
        }).dropna()

        # Converte para dict (o último valor encontrado para uma chave prevalece)
        # Invertemos a ordem se quisermos que o primeiro prevaleça,
        # mas aqui seguimos o comportamento de valor_corresp que pega o primeiro hit.
        # Então percorremos em ordem reversa para o dict.update ou apenas pegamos o primeiro.
        for row in temp_df.itertuples(index=False):
            v_prev = int(row.v_prev)
            if v_prev not in mapa:
                mapa[v_prev] = int(row.v_curr)
    return mapa


# ---------------------------------------------------------------------------
# Camada 1: matching por CD_CONTA CVM (prioritário)
# ---------------------------------------------------------------------------
def valor_corresp_por_conta(
    label: str | None,
    abas: Dict[str, pd.DataFrame],
    atual: str,
) -> int | None:
    """
    Dado o rótulo configurado no Spread, resolve os CD_CONTA equivalentes
    via MappingRegistry e soma os valores da coluna `atual` nas abas da Origem.

    Retorna None se:
    - label não está no mapa
    - nenhuma aba tem a coluna 'Codigo Conta'
    - soma resulta em zero (não gravamos zeros via Camada 1 — Camada 2 decide)
    """
    codes = layer1_codes_for_label(label)
    if not codes:
        return None
    for df in abas.values():
        col_codigo = next(
            (c for c in df.columns if str(c).strip().lower() == "codigo conta"),
            None,
        )
        if col_codigo is None or atual not in df.columns:
            continue
        mask = df[col_codigo].astype(str).isin(codes)
        if mask.any():
            total = sum(normaliza_num(v) or 0 for v in df.loc[mask, atual])
            if total != 0:
                return total
    return None


def atualizar_ws(
    ws,
    get_val: Callable[[int, int], object],
    set_val: Callable[[int, int, object], None],
    abas: Dict[str, pd.DataFrame],
    src_idx: int,
    dst_idx: int,
    atual: str,
    ant: str,
    start_row: int,
) -> tuple[List[int], Set[int], Set[int]]:
    """
    Copia e ajusta valores/fórmulas da coluna origem→destino EXCETO
    nas linhas especiais configuradas no SpreadSchema para DFC / DMPL.
    Retorna (skipped_rows, skipped_vals, used_vals).
    """
    c_src, c_dst = src_idx + 1, dst_idx + 1
    delta = c_dst - c_src
    skipped_rows: List[int] = []
    skipped_vals: Set[int] = set()
    used_vals: Set[int] = set()
    label_col = label_column_1based()
    schema_skip_rows = skip_rows()

    # Cache para Camada 2
    mapa_corresp = criar_mapa_corresp(abas, ant, atual)

    num_pat = re.compile(r"[-+]?\d[\d\.,]*")
    try:
        max_row = ws.max_row
    except AttributeError:
        max_row = ws.cells.last_cell.row

    empty_streak = 0
    r = start_row
    while empty_streak < 30 and r <= max_row:
        # pula as linhas de DFC/DMPL
        if r in schema_skip_rows:
            r += 1
            continue

        v = get_val(r, c_src)
        if v in (None, ""):
            empty_streak += 1
            r += 1
            continue
        empty_streak = 0

        wrote = False
        destino = v

        label = get_val(r, label_col)

        if isinstance(v, str) and v.startswith("="):
            if not re.search(r"[A-Za-z]", v[1:]):
                # soma/subtração de literais
                # ── Camada 2: substitui cada literal pelo valor correspondente ──
                def lit_repl(m: re.Match) -> str:
                    tok = m.group(0)
                    n0 = normaliza_num(tok.lstrip("+-"))
                    n1 = mapa_corresp.get(n0)
                    if n1 is not None:
                        used_vals.add(n1)
                        sign = tok[0] if tok[0] in "+-" else ""
                        return f"{sign}{abs(n1)}"
                    return tok

                destino_temp = "=" + num_pat.sub(lit_repl, v[1:])
                if destino_temp != v:
                    destino, wrote = destino_temp, True
                else:
                    # ── Camada 1: fallback por CD_CONTA se Camada 2 não achou nada ──
                    novo = valor_corresp_por_conta(label, abas, atual)
                    if novo is not None:
                        destino, wrote = novo, True
                        used_vals.add(novo)
            else:
                # fórmula complexa
                mp = lambda n: mapa_corresp.get(n)
                destino = adjust_complex_formula(v, delta, mp, used_vals)
                wrote = destino != v

        elif (n := normaliza_num(v)) is not None:
            # ── Camada 2: matching por valor numérico ──
            novo = mapa_corresp.get(n)
            # ── Camada 1: fallback por CD_CONTA CVM ──
            if novo is None:
                novo = valor_corresp_por_conta(label, abas, atual)
            if novo is not None:
                destino, wrote = novo, True
                used_vals.add(novo)
        else:
            wrote = True

        try:
            set_val(r, c_dst, destino)
        except Exception:
            wrote = False

        if not wrote and normaliza_num(v) not in (None, 0):
            skipped_rows.append(r)
            skipped_vals.add(normaliza_num(v) or 0)

        r += 1

    return skipped_rows, skipped_vals, used_vals


def coletar_vals_do_spread(
    spread_path: Path, dst_idx: int, start_row: int
) -> Set[int]:
    wb = load_workbook(spread_path, data_only=True)
    sheet_name = spread_sheet_name()
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    vals, empty = set(), 0
    r = start_row
    while empty < 30 and r <= ws.max_row:
        raw = ws.cell(r, dst_idx + 1).value
        if raw in (None, ""):
            empty += 1
        else:
            empty = 0
            n = normaliza_num(raw)
            if n is not None:
                vals.add(n)
        r += 1
    return vals
