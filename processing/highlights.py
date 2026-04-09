# processing/highlights.py — Destaque visual de células inseridas e novas
# Extraído de app_spread.py

from __future__ import annotations

from pathlib import Path
from typing import Set

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from core.utils import normaliza_num, XLWINGS

if XLWINGS:
    import xlwings as xw


def destacar_inseridos(
    orig_tratada: Path, used_vals: Set[int],
    atual: str, prefer_xlwings: bool = True
) -> None:
    if not used_vals:
        return
    if prefer_xlwings and XLWINGS:
        try:
            wb = xw.Book(str(orig_tratada))
            for sht in wb.sheets:
                hdrs = sht.range("A1").expand("right").value
                hdrs = hdrs if isinstance(hdrs, list) else [hdrs]
                cols = [i + 1 for i, h in enumerate(hdrs) if str(h).strip() == atual]
                last = sht.cells.last_cell.row
                for c in cols:
                    vals = sht.range((2, c), (last, c)).value
                    vals = vals if isinstance(vals, list) else [vals]
                    for i, v in enumerate(vals, start=2):
                        if normaliza_num(v) in used_vals:
                            cell = sht.cells(i, c)
                            cell.color = (204, 255, 204)
                            cell.api.Font.Bold = True
            wb.save()
            return
        except:
            pass
    wb = load_workbook(orig_tratada, keep_vba=orig_tratada.suffix.lower() == ".xlsm")
    fill, bold = PatternFill("solid", fgColor="CCFFCC"), Font(bold=True)
    for ws in wb.worksheets:
        cols = [c.column for c in ws[1] if str(c.value).strip() == atual]
        for row in ws.iter_rows(min_row=2):
            for c in cols:
                cell = row[c - 1]
                if normaliza_num(cell.value) in used_vals:
                    cell.fill, cell.font = fill, bold
    wb.save(orig_tratada)


def destacar_novos(
    orig_tratada: Path, prev: str, atual: str,
    prefer_xlwings: bool = True
) -> None:
    rgb_fill = PatternFill("solid", fgColor="99CCFF")
    bold = Font(bold=True)
    if prefer_xlwings and XLWINGS:
        try:
            for bk in xw.books:
                if Path(bk.fullname).resolve() == orig_tratada.resolve():
                    wb = bk
                    break
            else:
                wb = xw.Book(str(orig_tratada))
            for sht in wb.sheets:
                hdrs = sht.range("A1").expand("right").value
                hdrs = hdrs if isinstance(hdrs, list) else [hdrs]
                if prev not in hdrs or atual not in hdrs:
                    continue
                c_prev = hdrs.index(prev) + 1
                c_atual = hdrs.index(atual) + 1
                last = sht.cells.last_cell.row
                vp = sht.range((2, c_prev), (last, c_prev)).value
                va = sht.range((2, c_atual), (last, c_atual)).value
                vp = vp if isinstance(vp, list) else [vp]
                va = va if isinstance(va, list) else [va]
                for i, (pv, av) in enumerate(zip(vp, va), start=2):
                    if normaliza_num(pv) == 0 and normaliza_num(av) not in (None, 0):
                        cell = sht.cells(i, c_atual)
                        cell.color = (153, 204, 255)
                        cell.api.Font.Bold = True
            wb.save()
            return
        except:
            pass
    wb = load_workbook(orig_tratada, keep_vba=orig_tratada.suffix.lower() == ".xlsm")
    for ws in wb.worksheets:
        headers = {cell.value: cell.column for cell in ws[1]}
        c_prev = headers.get(prev)
        c_atual = headers.get(atual)
        if not c_prev or not c_atual:
            continue
        for row in ws.iter_rows(min_row=2):
            pv = row[c_prev - 1].value
            av = row[c_atual - 1].value
            if normaliza_num(pv) == 0 and normaliza_num(av) not in (None, 0):
                row[c_atual - 1].fill = rgb_fill
                row[c_atual - 1].font = bold
    wb.save(orig_tratada)
