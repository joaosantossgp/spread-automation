# processing/origin.py — Preparação do arquivo-origem (tratamento de abas)
# Extraído de app_spread.py

from __future__ import annotations

from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl import load_workbook


def prepara_origem(
    path: Path,
    tipo: str,
    atual: str,
    ant: str,
    ant2: str,
    is_trim: bool,
    out_dir: Path | None,
) -> Path:
    dst_dir = out_dir or path.parent
    dst_dir.mkdir(parents=True, exist_ok=True)
    dst_path = dst_dir / f"{path.stem}_tratado{path.suffix}"

    chapa = "Cons" if tipo == "consolidado" else "Ind"
    aba_dm = f"DF {chapa} DMPL {'Atual' if is_trim else 'Ultimo'}"
    sheet_map = {
        "consolidado": {
            "DF Cons Ativo": "cons ativos",
            "DF Cons Passivo": "cons passivos",
            "DF Cons Resultado Periodo": "cons DRE",
            "DF Cons Fluxo de Caixa": "cons DFC",
            aba_dm: "cons DMPL",
        },
        "individual": {
            "DF Ind Ativo": "ind ativos",
            "DF Ind Passivo": "ind passivos",
            "DF Ind Resultado Periodo": "ind DRE",
            "DF Ind Fluxo de Caixa": "ind DFC",
            aba_dm: "ind DMPL",
        },
    }[tipo]

    H_ANO = (
        "valor ultimo exercicio",
        "valor penultimo exercicio",
        "valor antepenultimo exercicio",
    )
    H_TRI_AP = ("valor trimestre atual", "valor exercicio anterior")
    H_TRI_RES = (
        "valor acumulado atual exercicio",
        "valor acumulado exercicio anterior",
    )

    def ren_factory(sheet_orig: str) -> Callable[[str], str]:
        low = sheet_orig.lower()
        is_ap = any(k in low for k in ("ativo", "passivo"))
        is_res = "resultado" in low or (is_trim and "fluxo" in low)

        def ren(col: str) -> str:
            c = col.lower().strip()
            if is_trim and is_ap:
                if c.startswith(H_TRI_AP[0]):
                    return atual
                if c.startswith(H_TRI_AP[1]):
                    return ant
            if is_trim and is_res:
                if c.startswith(H_TRI_RES[0]):
                    return atual
                if c.startswith(H_TRI_RES[1]):
                    return ant
            if c.startswith(H_ANO[0]):
                return atual
            if c.startswith(H_ANO[1]):
                return ant
            if c.startswith(H_ANO[2]):
                return ant2
            return col

        return ren

    engine = "openpyxl" if path.suffix.lower() in (".xlsx", ".xlsm") else None
    xls = pd.ExcelFile(path, engine=engine)
    with pd.ExcelWriter(dst_path, engine="openpyxl") as wr:
        for orig, novo in sheet_map.items():
            if orig not in xls.sheet_names:
                continue
            df = pd.read_excel(xls, sheet_name=orig, engine=engine)
            df = df.rename(columns=ren_factory(orig))
            df.to_excel(wr, sheet_name=novo, index=False)

    wb = load_workbook(dst_path)
    wb[wb.sheetnames[0]].sheet_state = "visible"
    wb.save(dst_path)
    return dst_path
