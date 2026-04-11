"""Structured spread slot detection for Mode 1A style flows."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from core.utils import col_txt_to_idx
from processing.runtime_bridge import (
    data_columns,
    has_next_data_column,
    next_data_column,
    spread_schema,
    spread_start_row,
)

SlotDetectionStatus = Literal["ready", "empty-grid", "no-slot-available"]

_SCAN_WINDOW_ROWS = 50


@dataclass(frozen=True)
class SpreadSlotDetection:
    """Normalized result for spread slot auto-detection."""

    quarterly: bool
    source_column: str | None
    destination_column: str | None
    occupied_columns: tuple[str, ...]
    status: SlotDetectionStatus
    message: str

    @property
    def has_available_slot(self) -> bool:
        return self.status == "ready"

    def as_pair(self) -> tuple[str, str]:
        if self.status == "empty-grid":
            raise SpreadGridEmptyError(self.message, detection=self)
        if not self.has_available_slot or self.source_column is None or self.destination_column is None:
            raise NoSpreadSlotAvailableError(self.message, detection=self)
        return self.source_column, self.destination_column


class SpreadSlotDetectionError(ValueError):
    """Base error raised by normalized spread slot detection."""

    def __init__(self, message: str, *, detection: SpreadSlotDetection):
        super().__init__(message)
        self.detection = detection


class SpreadGridEmptyError(SpreadSlotDetectionError):
    """Raised when no configured source column contains data."""


class NoSpreadSlotAvailableError(SpreadSlotDetectionError):
    """Raised when the configured spread grid has no writable slot left."""


def detect_mode1a_slot(
    spread_path: str | Path,
    period: str,
    *,
    start_row: int | None = None,
) -> SpreadSlotDetection:
    """Resolve source/destination columns for a Mode 1A period."""

    return detect_spread_slot(
        spread_path=spread_path,
        start_row=start_row,
        quarterly=_is_quarterly_period(period),
    )


def detect_spread_slot(
    spread_path: str | Path,
    *,
    start_row: int | None = None,
    quarterly: bool = False,
) -> SpreadSlotDetection:
    """Inspect the configured spread grid and return a normalized detection result."""

    path = Path(spread_path)
    schema = spread_schema()
    row_start = spread_start_row(start_row)

    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[schema.sheet_name] if schema.sheet_name in wb.sheetnames else wb.active
        annual_occupied = tuple(
            column
            for column in data_columns()
            if _column_has_values(ws, column, row_start)
        )
        quarterly_column = schema.columns.quarterly
        quarterly_occupied = _column_has_values(ws, quarterly_column, row_start)
    finally:
        wb.close()

    if quarterly:
        return _resolve_quarterly_slot(
            annual_occupied=annual_occupied,
            quarterly_column=quarterly_column,
            quarterly_occupied=quarterly_occupied,
        )

    return _resolve_annual_slot(annual_occupied)


def _resolve_annual_slot(annual_occupied: tuple[str, ...]) -> SpreadSlotDetection:
    if not annual_occupied:
        return SpreadSlotDetection(
            quarterly=False,
            source_column=None,
            destination_column=None,
            occupied_columns=(),
            status="empty-grid",
            message="Spread seems empty: none of the configured annual data columns contain values.",
        )

    source_column = annual_occupied[-1]
    if not has_next_data_column(source_column):
        return SpreadSlotDetection(
            quarterly=False,
            source_column=source_column,
            destination_column=None,
            occupied_columns=annual_occupied,
            status="no-slot-available",
            message=(
                f"Annual grid is full: latest populated column {source_column} has no next annual slot."
            ),
        )

    destination_column = next_data_column(source_column)
    return SpreadSlotDetection(
        quarterly=False,
        source_column=source_column,
        destination_column=destination_column,
        occupied_columns=annual_occupied,
        status="ready",
        message=(
            f"Detected annual source column {source_column} and destination column "
            f"{destination_column}."
        ),
    )


def _resolve_quarterly_slot(
    *,
    annual_occupied: tuple[str, ...],
    quarterly_column: str,
    quarterly_occupied: bool,
) -> SpreadSlotDetection:
    if quarterly_occupied:
        return SpreadSlotDetection(
            quarterly=True,
            source_column=annual_occupied[-1] if annual_occupied else None,
            destination_column=None,
            occupied_columns=annual_occupied + (quarterly_column,),
            status="no-slot-available",
            message=(
                f"Quarterly grid is full: destination column {quarterly_column} already contains data."
            ),
        )

    if not annual_occupied:
        return SpreadSlotDetection(
            quarterly=True,
            source_column=None,
            destination_column=None,
            occupied_columns=(),
            status="empty-grid",
            message="Spread seems empty: no populated annual column is available as a quarterly source.",
        )

    source_column = annual_occupied[-1]
    return SpreadSlotDetection(
        quarterly=True,
        source_column=source_column,
        destination_column=quarterly_column,
        occupied_columns=annual_occupied,
        status="ready",
        message=(
            f"Detected quarterly source column {source_column} and destination column "
            f"{quarterly_column}."
        ),
    )


def _column_has_values(ws, column: str, start_row: int) -> bool:
    column_index = col_txt_to_idx(column) + 1
    end_row = min(start_row + _SCAN_WINDOW_ROWS, ws.max_row + 1)
    for row_number in range(start_row, end_row):
        value = ws.cell(row_number, column_index).value
        if value is not None and value != "":
            return True
    return False


def _is_quarterly_period(period: str) -> bool:
    return "T" in str(period).strip().upper()


__all__ = [
    "NoSpreadSlotAvailableError",
    "SpreadGridEmptyError",
    "SpreadSlotDetection",
    "SpreadSlotDetectionError",
    "detect_mode1a_slot",
    "detect_spread_slot",
]
